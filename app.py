import time
import cv2
import numpy as np
import streamlit as st
from PIL import Image
import pandas as pd
import io
import os

# For converting HTML table to Excel with merged cells
from bs4 import BeautifulSoup
from openpyxl import Workbook

# Import recognition engines and helper functions
from lineless_table_rec import LinelessTableRecognition
from rapid_table import RapidTable, RapidTableInput
from rapid_table.main import ModelType
from rapidocr_onnxruntime import RapidOCR
from table_cls import TableCls
from wired_table_rec import WiredTableRecognition
from utils import plot_rec_box, LoadImage, format_html, box_4_2_poly_to_box_4_1

# -----------------------------------------------------------------------------
# Advanced Settings (Sidebar)
# -----------------------------------------------------------------------------
st.sidebar.header("Advanced Settings")

# New option: Extraction Mode
extraction_mode = st.sidebar.selectbox("Select Extraction Mode", ["Table Extraction", "Text Extraction"])

table_engine_type = st.sidebar.selectbox(
    "Select Recognition Table Engine",
    ["auto",
     "RapidTable(SLANet)",
     "RapidTable(SLANet-plus)",
     "wired_table_v2",
     "wired_table_v1",
     "lineless_table"],
    index=0
)
small_box_cut_enhance = st.sidebar.checkbox("Box Cutting Enhancement", value=True)
char_ocr = st.sidebar.checkbox("Character-level OCR", value=False)
rotated_fix = st.sidebar.checkbox("Table Rotate Rec Enhancement", value=False)
col_threshold = st.sidebar.slider("Column threshold (determine same col)", 5, 100, 15, step=5)
row_threshold = st.sidebar.slider("Row threshold (determine same row)", 5, 100, 10, step=5)

# -----------------------------------------------------------------------------
# Main Title and File Uploader (Main Body)
# -----------------------------------------------------------------------------
st.title("Extract Data from Image")

# First, provide an uploader in the main body.
uploaded_file = st.file_uploader("Upload an image", type=["jpg", "jpeg", "png"])

# If no image is uploaded, show sample images (thumbnails) for selection.
if uploaded_file is None and "selected_sample" not in st.session_state:
    st.info("No image uploaded. Please select one of the sample images below.")
    sample_folder = "sample"
    sample_files = [f for f in os.listdir(sample_folder) if f.lower().endswith((".jpg", ".jpeg", ".png"))]
    if sample_files:
        cols = st.columns(len(sample_files))
        for i, filename in enumerate(sample_files):
            filepath = os.path.join(sample_folder, filename)
            try:
                sample_img = Image.open(filepath).convert("RGB")
            except Exception as e:
                st.error(f"Error loading sample image {filename}: {e}")
                continue
            with cols[i]:
                st.image(sample_img, caption=filename, use_container_width=True)
                if st.button("Select", key=filepath):
                    st.session_state.selected_sample = filepath
    else:
        st.info("No sample images found in the sample folder.")

# Decide which image to use: uploaded file takes precedence over sample.
if uploaded_file is not None:
    try:
        input_img = Image.open(uploaded_file).convert("RGB")
    except Exception as e:
        st.error("Error loading uploaded image: " + str(e))
elif "selected_sample" in st.session_state:
    try:
        input_img = Image.open(st.session_state.selected_sample).convert("RGB")
    except Exception as e:
        st.error("Error loading sample image: " + str(e))
else:
    input_img = None

# Display the chosen image (if available)
if input_img is not None:
    st.image(input_img, caption="Selected Image", use_container_width=True)

# -----------------------------------------------------------------------------
# Cache Heavy Engine Loading
# -----------------------------------------------------------------------------
@st.cache_resource(show_spinner=False)
def load_engines():
    # Local model paths for OCR
    det_model_dir = {"mobile_det": "models/ocr/ch_PP-OCRv4_det_infer.onnx"}
    rec_model_dir = {"mobile_rec": "models/ocr/ch_PP-OCRv4_rec_infer.onnx"}
    
    # Local model path for table recognition (your SLANet model)
    table_rec_model_path = "models/table_rec/ch_ppstructure_mobile_v2_SLANet.onnx"
    
    # Initialize RapidTable engines (using your local SLANet model)
    rapid_table_engine = RapidTable(
        RapidTableInput(model_type=ModelType.PPSTRUCTURE_ZH.value, model_path=table_rec_model_path)
    )
    SLANet_plus_table_Engine = RapidTable(
        RapidTableInput(model_type=ModelType.SLANETPLUS.value, model_path=table_rec_model_path)
    )
    # Wired and Lineless engines
    wired_table_engine_v1 = WiredTableRecognition(version="v1")
    wired_table_engine_v2 = WiredTableRecognition(version="v2")
    lineless_table_engine = LinelessTableRecognition()
    table_cls = TableCls()
    
    # Build a dictionary of OCR engines using the local OCR model files.
    ocr_engine_dict = {}
    for det_model in det_model_dir.keys():
        for rec_model in rec_model_dir.keys():
            key = f"{det_model}_{rec_model}"
            ocr_engine_dict[key] = RapidOCR(
                det_model_path=det_model_dir[det_model],
                rec_model_path=rec_model_dir[rec_model]
            )
    
    return {
        "rapid_table_engine": rapid_table_engine,
        "SLANet_plus_table_Engine": SLANet_plus_table_Engine,
        "wired_table_engine_v1": wired_table_engine_v1,
        "wired_table_engine_v2": wired_table_engine_v2,
        "lineless_table_engine": lineless_table_engine,
        "table_cls": table_cls,
        "ocr_engine_dict": ocr_engine_dict,
    }

engines = load_engines()

# -----------------------------------------------------------------------------
# Helper Functions for OCR/Recognition
# -----------------------------------------------------------------------------
def trans_char_ocr_res(ocr_res):
    """Transform OCR results for character-level recognition."""
    word_result = []
    for res in ocr_res:
        score = res[2]
        for word_box, word in zip(res[3], res[4]):
            word_result.append([word_box, word, score])
    return word_result

def select_ocr_model(det_model, rec_model):
    """Return the OCR engine given model keys."""
    return engines["ocr_engine_dict"][f"{det_model}_{rec_model}"]

def select_table_model(img, table_engine_type, det_model, rec_model):
    """Select and return a table recognition engine based on the chosen type."""
    if table_engine_type == "RapidTable(SLANet)":
        return engines["rapid_table_engine"], table_engine_type
    elif table_engine_type == "RapidTable(SLANet-plus)":
        return engines["SLANet_plus_table_Engine"], table_engine_type
    elif table_engine_type == "wired_table_v1":
        return engines["wired_table_engine_v1"], table_engine_type
    elif table_engine_type == "wired_table_v2":
        return engines["wired_table_engine_v2"], table_engine_type
    elif table_engine_type == "lineless_table":
        return engines["lineless_table_engine"], table_engine_type
    elif table_engine_type == "auto":
        cls, _ = engines["table_cls"](img)
        if cls == 'wired':
            return engines["wired_table_engine_v2"], "wired_table_v2"
        return engines["lineless_table_engine"], "lineless_table"

def process_image(img_input, small_box_cut_enhance, table_engine_type, char_ocr, rotated_fix, col_threshold, row_threshold):
    """Process image for table extraction."""
    det_model = "mobile_det"
    rec_model = "mobile_rec"
    img_loader = LoadImage()
    img = img_loader(img_input)
    start = time.time()
    
    table_engine, table_type = select_table_model(img, table_engine_type, det_model, rec_model)
    ocr_engine = select_ocr_model(det_model, rec_model)
    
    ocr_res, ocr_infer_elapse = ocr_engine(img, return_word_box=char_ocr)
    det_cost, cls_cost, rec_cost = ocr_infer_elapse
    if char_ocr:
        ocr_res = trans_char_ocr_res(ocr_res)
    ocr_boxes = [box_4_2_poly_to_box_4_1(ori_ocr[0]) for ori_ocr in ocr_res]
    
    if isinstance(table_engine, RapidTable):
        table_results = table_engine(img, ocr_res)
        html, polygons, table_rec_elapse = table_results.pred_html, table_results.cell_bboxes, table_results.elapse
        # Adjust polygon format to (x1, y1, x2, y2)
        polygons = [[p[0], p[1], p[4], p[5]] for p in polygons]
    elif isinstance(table_engine, (WiredTableRecognition, LinelessTableRecognition)):
        html, table_rec_elapse, polygons, _, ocr_res = table_engine(
            img, ocr_result=ocr_res,
            enhance_box_line=small_box_cut_enhance,
            rotated_fix=rotated_fix,
            col_threshold=col_threshold,
            row_threshold=row_threshold
        )
    
    sum_elapse = time.time() - start
    all_elapse = (
        f"- table_type: {table_type}\n"
        f"- table all cost: {sum_elapse:.5f}\n"
        f"- table rec cost: {table_rec_elapse:.5f}\n"
        f"- ocr cost: {det_cost + cls_cost + rec_cost:.5f}"
    )
    
    # Convert image to RGB for display
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    table_boxes_img = plot_rec_box(img.copy(), polygons)
    ocr_boxes_img = plot_rec_box(img.copy(), ocr_boxes)
    complete_html = format_html(html)
    return complete_html, table_boxes_img, ocr_boxes_img, all_elapse

def process_text_image(img_input, char_ocr):
    """Process image for text extraction using only OCR."""
    det_model = "mobile_det"
    rec_model = "mobile_rec"
    img_loader = LoadImage()
    img = img_loader(img_input)
    start = time.time()
    
    ocr_engine = select_ocr_model(det_model, rec_model)
    ocr_res, ocr_infer_elapse = ocr_engine(img, return_word_box=char_ocr)
    det_cost, cls_cost, rec_cost = ocr_infer_elapse
    if char_ocr:
        ocr_res = trans_char_ocr_res(ocr_res)
    ocr_boxes = [box_4_2_poly_to_box_4_1(ori_ocr[0]) for ori_ocr in ocr_res]
    
    # Extract recognized text. Assumes each OCR result's second element is the text.
    if char_ocr:
        texts = [word for (box, word, score) in ocr_res]
    else:
        texts = [res[1] for res in ocr_res]
    full_text = "\n".join(texts)
    
    sum_elapse = time.time() - start
    all_elapse = (
        f"- text extraction cost: {sum_elapse:.5f}\n"
        f"- ocr cost: {det_cost + cls_cost + rec_cost:.5f}"
    )
    
    ocr_boxes_img = plot_rec_box(img.copy(), ocr_boxes)
    return full_text, ocr_boxes_img, all_elapse

# -----------------------------------------------------------------------------
# Helper Function: Convert HTML table to Excel (preserving merged cells)
# -----------------------------------------------------------------------------
def html_table_to_excel_bytes(html):
    """
    Parse the HTML table using BeautifulSoup and create an Excel workbook with merged cells.
    Returns the Excel file as bytes.
    """
    soup = BeautifulSoup(html, "lxml")
    table = soup.find("table")
    if table is None:
        raise ValueError("No table found in HTML.")
    
    wb = Workbook()
    ws = wb.active

    current_row = 1
    # Track occupied cells for merged regions.
    occupied = {}
    for tr in table.find_all("tr"):
        current_col = 1
        for cell in tr.find_all(["td", "th"]):
            # Move to next free column
            while (current_row, current_col) in occupied:
                current_col += 1

            text = cell.get_text(strip=True)
            colspan = int(cell.get("colspan", 1))
            rowspan = int(cell.get("rowspan", 1))
            
            ws.cell(row=current_row, column=current_col, value=text)
            
            # Mark cells as occupied and merge if needed.
            if colspan > 1 or rowspan > 1:
                start_cell = ws.cell(row=current_row, column=current_col).coordinate
                end_cell = ws.cell(row=current_row + rowspan - 1, column=current_col + colspan - 1).coordinate
                ws.merge_cells(f"{start_cell}:{end_cell}")
                for i in range(current_row, current_row + rowspan):
                    for j in range(current_col, current_col + colspan):
                        occupied[(i, j)] = True
            else:
                occupied[(current_row, current_col)] = True

            current_col += 1
        current_row += 1

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# -----------------------------------------------------------------------------
# Main App Workflow
# -----------------------------------------------------------------------------
if input_img is not None:
    if st.button("Run Recognition"):
        if extraction_mode == "Table Extraction":
            with st.spinner("Processing image for table extraction..."):
                complete_html, table_boxes_img, ocr_boxes_img, all_elapse = process_image(
                    input_img, small_box_cut_enhance, table_engine_type,
                    char_ocr, rotated_fix, col_threshold, row_threshold
                )
            st.success("Processing complete!")
            st.markdown("### Recognized Table (HTML)")
            st.markdown(complete_html, unsafe_allow_html=True)
            st.markdown("### Table Recognition Boxes")
            st.image(table_boxes_img, use_container_width=True)
            st.markdown("### OCR Boxes")
            st.image(ocr_boxes_img, use_container_width=True)
            st.markdown("### Elapsed Time")
            st.text(all_elapse)
            
            # Convert the HTML table to an Excel file preserving merged cells.
            try:
                excel_bytes = html_table_to_excel_bytes(complete_html)
                st.download_button(
                    label="Download Excel (.xlsx)",
                    data=excel_bytes,
                    file_name="table.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                # Optionally, also show a flat DataFrame view parsed from the HTML.
                df_list = pd.read_html(complete_html)
                if df_list:
                    df = df_list[0]
                    st.markdown("### Extracted Table Data (Flat View)")
                    st.dataframe(df, use_container_width=True)
            except Exception as e:
                st.error("Failed to convert HTML table to Excel: " + str(e))
        elif extraction_mode == "Text Extraction":
            with st.spinner("Processing image for text extraction..."):
                full_text, ocr_boxes_img, all_elapse = process_text_image(input_img, char_ocr)
            st.success("Processing complete!")
            st.markdown("### Extracted Text")
            st.text_area("OCR Result", full_text, height=200)
            st.download_button(
                label="Download Text",
                data=full_text,
                file_name="extracted_text.txt",
                mime="text/plain"
            )
            st.markdown("### OCR Boxes")
            st.image(ocr_boxes_img, use_container_width=True)
            st.markdown("### Elapsed Time")
            st.text(all_elapse)
else:
    st.info("Please upload an image or select a sample image to begin.")
